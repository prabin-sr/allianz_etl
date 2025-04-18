import zipfile
from pathlib import Path
from unittest.mock import AsyncMock, patch

import pytest
from openpyxl import load_workbook

from etl.core.scraper import DataScraper
from tests.mock_data.mock_data import (
    BASE_URL,
    get_analytics_sheet_name,
    get_excel_sheet_names,
    get_mocked_base_url_html,
    get_mocked_pagination_url_html,
    get_mocket_analytical_data,
    get_zipped_files_list,
)


@pytest.fixture
def mocked_scraper():
    with patch.object(DataScraper, "__init__", lambda x: None):
        scraper = DataScraper()
        scraper._base_url = BASE_URL
        scraper._media_dir = Path.joinpath(
            Path(__file__).parent.resolve(), "mock_media"
        )
        scraper._zip_file_name = "mocked.zip"
        scraper._excel_file_name = "mocked.xlsx"
        scraper._extracted_data_headers = [
            "Team Name",
            "Year",
            "Wins",
            "Losses",
            "OT Losses",
            "Win %",
            "Goals For (GF)",
            "Goals Against (GA)",
            "+ / -",
        ]
        scraper._analytics_data_headers = [
            "Year",
            "Winner",
            "Winner Num. of Wins",
            "Loser",
            "Loser Num. of Wins",
        ]
        scraper._extracted_data_sheet_name = "NHL Stats 1990-2011"
        scraper._analytics_data_sheet_name = "Winner and Loser per Year"
        yield scraper


@pytest.mark.asyncio
async def test_main(mocked_scraper: DataScraper):
    base_url_html = get_mocked_base_url_html()

    pagination_url_html = get_mocked_pagination_url_html()

    with patch.object(
        DataScraper, "fetch_all_pages", new_callable=AsyncMock
    ) as mock_fetch:
        mock_fetch.side_effect = [
            base_url_html,
            pagination_url_html,
        ]

        await mocked_scraper()

        await mocked_scraper._zip_task
        assert mock_fetch.call_count == 2

        zip_path = Path.joinpath(
            mocked_scraper._media_dir, mocked_scraper._zip_file_name
        )

        with zipfile.ZipFile(zip_path, "r") as zipf:
            zip_file_list = zipf.namelist()
            assert zip_file_list == get_zipped_files_list()

        excel_path = Path.joinpath(
            mocked_scraper._media_dir, mocked_scraper._excel_file_name
        )

        workbook = load_workbook(excel_path)
        assert workbook.sheetnames == get_excel_sheet_names()

        sheet = workbook[get_analytics_sheet_name()]

        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(row)

        mocked_sheet_data = get_mocket_analytical_data()

        assert sheet_data == mocked_sheet_data
